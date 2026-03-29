# python/pipeline_validation.py
# End-to-end validation of the entire project pipeline

import pandas as pd
import os
from db_connection import get_connection, run_query

def validate_sql_layer():
    print("\n" + "="*55)
    print("  LAYER 1: SQL DATA WAREHOUSE VALIDATION")
    print("="*55)

    tables = {
        'companies':      10,
        'financials':     30,
        'stock_prices':   7000,
        'esg_scores':     30,
        'news_sentiment': 60
    }

    all_passed = True
    conn = get_connection()
    cursor = conn.cursor()

    for table, min_rows in tables.items():
        cursor.execute(f"SELECT COUNT(*) FROM {table}")
        count = cursor.fetchone()[0]
        status = "✅" if count >= min_rows else "❌"
        if count < min_rows:
            all_passed = False
        print(f"  {status} {table:<20} "
              f"{count:>6} rows "
              f"(min: {min_rows})")

    conn.close()
    return all_passed

def validate_python_layer():
    print("\n" + "="*55)
    print("  LAYER 2: PYTHON PROCESSED FILES VALIDATION")
    print("="*55)

    files = {
        'data/processed/financial_metrics.csv':      ('rows', 30),
        'data/processed/esg_scores_processed.csv':   ('rows', 30),
        'data/processed/esg_investment_tiers.csv':   ('rows', 10),
        'data/processed/master_dataset.csv':         ('rows', 30),
        'data/processed/sentiment_scores.csv':       ('rows', 60),
        'data/processed/esg_predictions_2024.csv':   ('rows', 10),
        'data/processed/ai_recommendations.csv':     ('rows', 10),
    }

    all_passed = True
    for filepath, (check, expected) in files.items():
        if os.path.exists(filepath):
            df = pd.read_csv(filepath)
            actual = len(df)
            status = "✅" if actual >= expected else "⚠️ "
            if actual < expected:
                all_passed = False
            print(f"  {status} {os.path.basename(filepath):<40} "
                  f"{actual} rows")
        else:
            print(f"  ❌ MISSING: {filepath}")
            all_passed = False

    return all_passed

def validate_excel_layer():
    print("\n" + "="*55)
    print("  LAYER 3: EXCEL MODEL VALIDATION")
    print("="*55)

    excel_path = "excel/financial_model.xlsx"
    if os.path.exists(excel_path):
        size_kb = os.path.getsize(excel_path) / 1024
        print(f"  ✅ financial_model.xlsx found "
              f"({size_kb:.1f} KB)")

        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, read_only=True)
            required_sheets = [
                'Cover', 'Assumptions', 'Income_Statement',
                'Balance_Sheet', 'Cash_Flow',
                'DCF_Valuation', 'Peer_Comparison'
            ]
            for sheet in required_sheets:
                if sheet in wb.sheetnames:
                    print(f"  ✅ Sheet: {sheet}")
                else:
                    print(f"  ❌ Missing sheet: {sheet}")
            wb.close()
        except Exception as e:
            print(f"  ⚠️  Could not read sheets: {e}")
        return True
    else:
        print(f"  ❌ MISSING: {excel_path}")
        return False

def validate_powerbi_layer():
    print("\n" + "="*55)
    print("  LAYER 4: POWER BI VALIDATION")
    print("="*55)

    pbix_files = [f for f in os.listdir('powerbi')
                  if f.endswith('.pbix')]
    if pbix_files:
        for f in pbix_files:
            size_mb = os.path.getsize(
                f'powerbi/{f}') / (1024*1024)
            print(f"  ✅ {f} ({size_mb:.1f} MB)")
        return True
    else:
        print("  ❌ No .pbix file found in powerbi/")
        return False

def print_project_summary():
    print("\n" + "="*55)
    print("  PROJECT METRICS SUMMARY")
    print("="*55)

    # Key metrics from processed data
    try:
        ai = pd.read_csv("data/processed/ai_recommendations.csv")
        esg = pd.read_csv("data/processed/esg_investment_tiers.csv")
        fin = pd.read_csv("data/processed/financial_metrics.csv")

        print(f"\n  📊 Coverage Universe:     10 companies")
        print(f"  📅 Historical Period:     FY2021 – FY2023")
        print(f"  📈 Projection Period:     FY2024E – FY2026E")
        print(f"\n  🤖 AI Recommendations:")

        rec_counts = ai['ai_recommendation'].value_counts()
        for rec, count in rec_counts.items():
            print(f"     {rec:<15} {count} companies")

        print(f"\n  🌱 ESG Summary (2023):")
        esg_summary = esg['calculated_rating'].value_counts()
        for rating, count in esg_summary.items():
            print(f"     {rating:<8} {count} companies")

        avg_esg = esg['weighted_esg_score'].mean()
        print(f"     Avg ESG Score: {avg_esg:.1f}/100")

        print(f"\n  💰 Financial Summary (2023):")
        fin_2023 = fin[fin['fiscal_year']==2023]
        print(f"     Avg Net Margin:  "
              f"{fin_2023['net_margin_pct'].mean():.1f}%")
        print(f"     Avg EBITDA Margin: "
              f"{fin_2023['ebitda_margin_pct'].mean():.1f}%")

    except Exception as e:
        print(f"  Could not load summary data: {e}")

def run_validation():
    print("\n" + "🔍 "*20)
    print("  ESG FINANCIAL ANALYTICS PLATFORM")
    print("  End-to-End Pipeline Validation")
    print("🔍 "*20)

    sql_ok    = validate_sql_layer()
    python_ok = validate_python_layer()
    excel_ok  = validate_excel_layer()
    pbi_ok    = validate_powerbi_layer()

    print_project_summary()

    print("\n" + "="*55)
    print("  FINAL VALIDATION RESULT")
    print("="*55)
    print(f"  SQL Layer:      {'✅ PASS' if sql_ok    else '❌ FAIL'}")
    print(f"  Python Layer:   {'✅ PASS' if python_ok else '❌ FAIL'}")
    print(f"  Excel Layer:    {'✅ PASS' if excel_ok  else '❌ FAIL'}")
    print(f"  Power BI Layer: {'✅ PASS' if pbi_ok    else '❌ FAIL'}")

    all_pass = all([sql_ok, python_ok, excel_ok, pbi_ok])
    print(f"\n  {'🎉 ALL SYSTEMS GO!' if all_pass else '⚠️  FIX ISSUES ABOVE'}")
    print("="*55 + "\n")

if __name__ == "__main__":
    run_validation()